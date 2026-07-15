from __future__ import annotations

import json
import pytest
from pathlib import Path

from core.spectra.static_coefficients import SpectrumCurve, _curve_at_elevation, resolve_spectrum_elevation
from core.spectra.response_spectrum_writer import _validate_formal_spectrum_curve, write_segmented_response_spectrum_mac
from core.spectra.selector import select_spectrum_points


def _curve(source: str) -> SpectrumCurve:
    return SpectrumCurve((1.0, 100.0), (0.1, 0.2), source)


def test_formal_spectrum_curve_integrity_gate_accepts_valid_curve() -> None:
    audit = _validate_formal_spectrum_curve(_curve("valid"), label="SL-1(XY)")

    assert audit["status"] == "pass"
    assert audit["frequency_max_hz"] == 100.0


@pytest.mark.parametrize(
    "curve, message",
    [
        (SpectrumCurve((1.0, 1.0, 100.0), (0.1, 0.2, 0.3), "duplicate"), "strictly increasing"),
        (SpectrumCurve((1.0, 100.0), (0.0, 0.0), "zero"), "all zero"),
        (SpectrumCurve((1.0, 50.0), (0.1, 0.2), "tail"), "100 Hz tail"),
    ],
)
def test_formal_spectrum_curve_integrity_gate_blocks_invalid_data(curve: SpectrumCurve, message: str) -> None:
    with pytest.raises(ValueError, match=message):
        _validate_formal_spectrum_curve(curve, label="SL-1(XY)")


def _curves_for_elevations(elevations: list[float]) -> dict[tuple[str, str, float, float], SpectrumCurve]:
    curves: dict[tuple[str, str, float, float], SpectrumCurve] = {}
    for level, damping in (("SL-1", 0.07), ("SL-2", 0.10)):
        for direction in ("X", "Y", "Z"):
            for elevation in elevations:
                curves[(level, direction, elevation, damping)] = _curve(f"{level}:{direction}:{elevation:g}")
    return curves


def _flat_spectrum_data(elevations: list[float]) -> dict:
    points = []
    for elevation in elevations:
        for frequency, acceleration in ((1.0, elevation), (2.0, elevation + 0.1)):
            points.append(
                {
                    "project_code": "P",
                    "building": "B",
                    "area": "A",
                    "level": "SL-1",
                    "direction": "X",
                    "elevation": elevation,
                    "damping": 0.07,
                    "frequency_hz": frequency,
                    "acceleration_g": acceleration,
                    "source_ref": f"B:{elevation:g}:{frequency:g}",
                }
            )
    return {"points": points}


def _add_segmented_section(sheet, label: str, scale: float) -> None:
    sheet.append([label])
    for frequency, base in ((0.1, 0.015), (3.927, 0.733), (4.094, 0.658), (5.482, 1.514), (100.0, 0.477)):
        sheet.append([None, frequency, base * scale, base * scale, base * scale, base * scale, base * scale, base * scale])


def _add_precision_control_sheet(workbook) -> None:
    sheet = workbook.create_sheet("精度控制")
    values = [50, 0, 33, 0, 15, 0, 3, 0, 0]
    for group_index in range(1, 9):
        row0 = 5 + ((group_index - 1) % 4) * 19
        col0 = 1 + ((group_index - 1) // 4) * 11
        for offset, value in enumerate(values):
            sheet.cell(row0 + offset, col0, value)


def _write_workbook_with_active_ansys_format(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    review = workbook.active
    review.title = "Review"
    ansys_lines = [
        "!SL-1(XY) 7%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,100.000",
        "SV, 0.07,  0.015,  0.733,  1.514,  0.182",
        "!SL-1(Z) 7%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,100.000",
        "SV, 0.07,  0.009,  0.502,  1.231,  0.117",
        "!SL-2(XY) 10%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000",
        "SV, 0.10,  0.043,  0.733,  1.514,  1.737,  0.477",
        "!SL-2(Z) 10%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000",
        "SV, 0.10,  0.024,  0.451,  1.231,  1.360,  0.337",
    ]
    for row_number, line in enumerate(ansys_lines, start=1):
        review.cell(row_number, 13, line)

    _add_precision_control_sheet(workbook)
    segmented = workbook.create_sheet("NR_1818")
    for label, scale in (
        ("8.45x", 1.0),
        ("8.45y", 1.1),
        ("8.45z", 0.8),
        ("8.45X", 1.2),
        ("8.45Y", 1.3),
        ("8.45Z", 0.9),
    ):
        _add_segmented_section(segmented, label, scale)
    workbook.save(path)
    workbook.close()


def _write_workbook_without_active_ansys_format(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "Review"
    _add_precision_control_sheet(workbook)
    segmented = workbook.create_sheet("NR_1818")
    for label, scale in (
        ("8.45x", 1.0),
        ("8.45y", 1.1),
        ("8.45z", 0.8),
        ("8.45X", 1.2),
        ("8.45Y", 1.3),
        ("8.45Z", 0.9),
    ):
        _add_segmented_section(segmented, label, scale)
    workbook.save(path)
    workbook.close()


def test_intake_850_uses_845_floor_within_lower_tolerance() -> None:
    curves = _curves_for_elevations([4.75, 8.45, 12.95])

    resolved = resolve_spectrum_elevation(
        curves,
        elevation=8.5,
        requirements=[("SL-1", "X", 0.07), ("SL-2", "Z", 0.10)],
    )

    assert resolved["mode"] == "lower_floor_within_0p1m_tolerance"
    assert resolved["requested_elevation"] == 8.5
    assert resolved["selected_elevation"] == 8.45
    assert resolved["minimum_allowed_elevation"] == pytest.approx(8.4)

    curve = _curve_at_elevation(curves, level="SL-1", direction="X", elevation=resolved["selected_elevation"], damping=0.07)
    assert curve.source_ref == "SL-1:X:8.45"


def test_true_between_floor_elevation_uses_next_upper_floor_without_interpolation() -> None:
    curves = _curves_for_elevations([4.75, 8.45, 12.95])

    resolved = resolve_spectrum_elevation(
        curves,
        elevation=10.0,
        requirements=[("SL-1", "X", 0.07), ("SL-2", "Z", 0.10)],
    )

    assert resolved["mode"] == "next_upper_floor"
    assert resolved["selected_elevation"] == 12.95
    curve = _curve_at_elevation(curves, level="SL-1", direction="X", elevation=12.95, damping=0.07)
    assert curve.source_ref == "SL-1:X:12.95"


def test_lower_floor_outside_tolerance_uses_next_upper_floor() -> None:
    curves = _curves_for_elevations([7.5, 13.5])

    resolved = resolve_spectrum_elevation(
        curves,
        elevation=8.0,
        requirements=[("SL-1", "X", 0.07), ("SL-2", "Z", 0.10)],
    )

    assert resolved["mode"] == "next_upper_floor"
    assert resolved["selected_elevation"] == 13.5
    assert resolved["minimum_allowed_elevation"] == pytest.approx(7.9)


def test_legacy_flat_selector_uses_same_no_interpolation_elevation_policy() -> None:
    query = {
        "project_code": "P",
        "building": "B",
        "area": "A",
        "level": "SL-1",
        "direction": "X",
        "elevation": 8.0,
        "damping": 0.07,
    }

    selection = select_spectrum_points(_flat_spectrum_data([7.5, 13.5]), query)

    assert selection["selection_mode"] == "next_upper_floor"
    assert selection["selected_elevation"] == 13.5
    assert selection["source_elevations"] == [13.5]
    assert [item["acceleration_g"] for item in selection["points"]] == [13.5, 13.6]


def test_out_of_range_elevation_is_blocked_instead_of_using_nearest() -> None:
    curves = _curves_for_elevations([4.75, 8.45, 12.95])

    with pytest.raises(ValueError, match="No common spectrum elevation >= requested elevation minus"):
        resolve_spectrum_elevation(
            curves,
            elevation=30.0,
            requirements=[("SL-1", "X", 0.07), ("SL-2", "Z", 0.10)],
        )

    with pytest.raises(ValueError, match="No exact spectrum elevation"):
        _curve_at_elevation(curves, level="SL-1", direction="X", elevation=30.0, damping=0.07)


def test_spectrum_macro_generation_can_auto_envelope_without_active_column_m_ansys_format(tmp_path: Path) -> None:
    workbook = tmp_path / "floor_spectrum_without_m.xlsx"
    _write_workbook_without_active_ansys_format(workbook)

    audit = write_segmented_response_spectrum_mac(
        workbook,
        tmp_path,
        project_code="1818",
        building="NR",
        elevation=8.5,
    )

    selection = json.loads((tmp_path / "spectrum_selection.json").read_text(encoding="utf-8"))
    assert audit["formal_spectrum_source_mode"] == "python_vba_envelope_replicator"
    assert selection["elevation_selection"]["mode"] == "python_vba_envelope_replicator"
    assert selection["elevation_selection"]["active_workbook_column_m_comparison"]["status"] == "not_applicable"
    assert (tmp_path / "ansys_spectrum_sl1.mac").exists()
    assert (tmp_path / "ansys_spectrum_sl2.mac").exists()


def test_active_column_m_ansys_format_is_used_verbatim_for_solve_macros(tmp_path: Path) -> None:
    workbook = tmp_path / "floor_spectrum.xlsx"
    _write_workbook_with_active_ansys_format(workbook)

    audit = write_segmented_response_spectrum_mac(
        workbook,
        tmp_path,
        project_code="1818",
        building="NR",
        elevation=8.5,
    )

    selection = json.loads((tmp_path / "spectrum_selection.json").read_text(encoding="utf-8"))
    sl2 = (tmp_path / "ansys_spectrum_sl2.mac").read_text(encoding="utf-8")
    zpa = (tmp_path / "ansys_zpa_parameters.mac").read_text(encoding="utf-8")
    assert audit["selected_elevation"] == 8.5
    assert selection["elevation_selection"]["mode"] == "python_vba_envelope_replicator"
    assert audit["active_workbook_column_m_comparison"]["status"] == "pass"
    assert audit["formal_spectrum_source_mode"] == "active_column_m_calibrated_precision_output"
    assert selection["elevation_selection"]["precision_control_override"]["status"] == "used"
    assert "active_column_m_calibrated_precision_output" in sl2
    assert "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000" in sl2
    assert "SV, 0.10,  0.043,  0.733,  1.514,  1.737,  0.477" in sl2
    assert "4.094" not in sl2
    assert "MMASS" not in sl2
    assert "pasx=static_factor*0.477*9.81" in zpa


def test_spectrum_macro_generation_falls_back_when_packaged_runtime_lacks_expat(tmp_path: Path, monkeypatch) -> None:
    workbook = tmp_path / "floor_spectrum_no_expat.xlsx"
    _write_workbook_with_active_ansys_format(workbook)

    import openpyxl

    def fail_load_workbook(*_args, **_kwargs):
        raise ImportError("No module named expat; use SimpleXMLTreeBuilder instead")

    monkeypatch.setattr(openpyxl, "load_workbook", fail_load_workbook)

    audit = write_segmented_response_spectrum_mac(
        workbook,
        tmp_path,
        project_code="1818",
        building="NR",
        elevation=8.5,
    )

    sl2 = (tmp_path / "ansys_spectrum_sl2.mac").read_text(encoding="utf-8")
    assert audit["active_workbook_column_m_comparison"]["status"] == "pass"
    assert audit["formal_spectrum_source_mode"] == "active_column_m_calibrated_precision_output"
    assert "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000" in sl2
