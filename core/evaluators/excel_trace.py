from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.evaluators.formula_registry import FORMULA_REGISTRY, todo_formulas


TODO_MARKERS = ("OFFSET", "MATCH", "INDIRECT", "VLOOKUP", "HLOOKUP", "INDEX")
CELL_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z0-9_\u4e00-\u9fff]+)?!?\$?[A-Z]{1,3}\$?\d+")
EVALUATION_WORKBOOK_TOKEN = "\u7ed3\u679c\u8bc4\u5b9a"


def _evaluation_workbooks(source_root: Path) -> list[Path]:
    return sorted(
        [
            path
            for path in source_root.rglob("*.xlsx")
            if EVALUATION_WORKBOOK_TOKEN in path.name and not path.name.startswith("~$")
        ],
        key=lambda item: item.name,
    )


def _formula_type(formula: str) -> str:
    upper = formula.upper()
    if "MIN(" in upper and any(token in upper for token in ("0.45", "0.37", "0.4", "0.33")):
        return "material_allowable"
    if "^2" in upper or "SQRT" in upper:
        return "interaction_or_equivalent"
    if "/" in upper:
        return "ratio"
    if any(marker in upper for marker in TODO_MARKERS):
        return "lookup_or_dynamic"
    return "other"


def _precedents(formula: str) -> list[str]:
    return sorted(set(match.group(0).replace("$", "") for match in CELL_REF_RE.finditer(formula)))


def extract_formula_traceability(
    source_root: Path | str = Path("source_materials"),
    output_path: Path | str = Path("docs/formula_traceability.md"),
) -> dict[str, Any]:
    source_root = Path(source_root)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    formulas: list[dict[str, Any]] = []
    manual_confirmation: list[dict[str, Any]] = []
    for workbook_path in _evaluation_workbooks(source_root):
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        values_workbook = load_workbook(workbook_path, data_only=True, read_only=False)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            values_sheet = values_workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value
                        record = {
                            "workbook": workbook_path.as_posix(),
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": formula,
                            "cached_value": values_sheet[cell.coordinate].value,
                            "precedents": _precedents(formula),
                            "formula_type": _formula_type(formula),
                        }
                        formulas.append(record)
                        if any(marker in formula.upper() for marker in TODO_MARKERS):
                            manual_confirmation.append(record)

    lines = [
        "# Formula Traceability",
        "",
        "Formula cells are extracted from authoritative evaluation workbooks without modifying source files.",
        "",
        f"- Formula cells found: {len(formulas)}",
        f"- Formulas requiring manual confirmation: {len(manual_confirmation)}",
        "",
        "## Formula Registry",
        "",
        "| Formula id | Status | Source ref | Description |",
        "| --- | --- | --- | --- |",
    ]
    for formula_id, record in FORMULA_REGISTRY.items():
        lines.append(f"| {formula_id} | {record.status} | {record.source_ref} | {record.description} |")

    lines.extend(
        [
            "",
            "## Formula Cell Inventory",
            "",
            "| Workbook | Sheet | Cell | Type | Formula | Cached value | Precedents |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for record in formulas:
        formula = str(record["formula"]).replace("|", "\\|")
        precedents = ", ".join(record["precedents"])
        lines.append(
            f"| {record['workbook']} | {record['sheet']} | {record['cell']} | {record['formula_type']} | `{formula}` | {record['cached_value']} | {precedents} |"
        )

    lines.extend(
        [
            "",
            "## Manual Confirmation List",
            "",
            "| Formula id or workbook cell | Reason |",
            "| --- | --- |",
        ]
    )
    for formula_id, record in todo_formulas().items():
        lines.append(f"| {formula_id} | {record.description} |")
    for record in manual_confirmation:
        formula = str(record["formula"]).replace("|", "\\|")
        lines.append(
            f"| {record['workbook']}:{record['sheet']}!{record['cell']} | Dynamic formula `{formula}` needs deterministic replication review. |"
        )

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {"formulas": formulas, "manual_confirmation": manual_confirmation}


if __name__ == "__main__":
    result = extract_formula_traceability()
    print(json.dumps({"formulas": len(result["formulas"]), "manual_confirmation": len(result["manual_confirmation"])}, ensure_ascii=False))
