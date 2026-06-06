from __future__ import annotations

import hashlib
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from core.intake.intake_excel_reader import _openpyxl_unavailable_due_to_expat, _read_xlsx_rows_without_expat


DAMPING_BY_LEVEL = {"SL-1": 0.07, "SL-2": 0.10}
DAMPING_COLUMNS = (0.02, 0.03, 0.04, 0.05, 0.07, 0.10)
DEFAULT_ALIAS_PATH = Path("data/spectra/building_aliases.json")
SPECTRUM_ELEVATION_LOWER_TOLERANCE_M = 0.1


@dataclass(frozen=True)
class SpectrumCurve:
    frequency_hz: tuple[float, ...]
    acceleration_g: tuple[float, ...]
    source_ref: str

    @property
    def peak_g(self) -> float:
        return max(self.acceleration_g) if self.acceleration_g else 0.0

    @property
    def zero_period_g(self) -> float:
        """Return the high-frequency tail used by source command streams as zero-period acceleration."""
        if not self.acceleration_g:
            return 0.0
        return self.acceleration_g[-1]

    def acceleration_at_hz(self, frequency_hz: float) -> float:
        """Return the spectrum acceleration at a requested frequency by log-frequency interpolation."""
        return _interp_log_frequency(self, frequency_hz)


@dataclass(frozen=True)
class SectionLabel:
    elevation_m: float
    direction: str
    level: str


def sha256_file(path: Path | str) -> str:
    path = Path(path)
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _normalise_text(value: Any) -> str:
    text = str(value or "").strip()
    table = str.maketrans({"，": ",", "（": "(", "）": ")", "：": ":"})
    return "".join(text.translate(table).split()).lower()


def _parse_section_label(value: Any) -> SectionLabel | None:
    if not isinstance(value, str):
        return None
    match = re.match(r"^\s*([+-]?\d+(?:\.\d+)?)\s*([xyzXYZ])\s*$", value)
    if not match:
        return None
    direction_token = match.group(2)
    return SectionLabel(
        elevation_m=float(match.group(1)),
        direction=direction_token.upper(),
        level="SL-1" if direction_token.islower() else "SL-2",
    )


def _sheet_has_segmented_spectrum(rows: list[tuple[Any, ...]]) -> bool:
    return any(_parse_section_label(row[0] if row else None) for row in rows)


def _normalise_rows(rows: list[Any]) -> list[tuple[Any, ...]]:
    return [tuple(row or ()) for row in rows]


def _read_workbook_rows_with_openpyxl(path: Path) -> list[tuple[str, list[tuple[Any, ...]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        return [
            (sheet.title, _normalise_rows(list(sheet.iter_rows(values_only=True))))
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _read_workbook_rows(path: Path) -> list[tuple[str, list[tuple[Any, ...]]]]:
    """Read XLSX/XLSM rows, falling back when packaged Python lacks expat.

    Some unit intranet computers run the PyInstaller bundle without the XML
    expat extension available to openpyxl.  The intake parser already has a
    small no-expat XLSX reader; reuse it for spectrum previews so deployment
    does not depend on that optional extension being present.
    """
    try:
        return _read_workbook_rows_with_openpyxl(path)
    except Exception as exc:
        if not _openpyxl_unavailable_due_to_expat(exc):
            raise
        return [(sheet_name, _normalise_rows(rows)) for sheet_name, rows in _read_xlsx_rows_without_expat(path)]


def _read_segmented_sheet(path: Path, sheet_name: str) -> dict[tuple[str, str, float, float], SpectrumCurve]:
    sheet_items = dict(_read_workbook_rows(path))
    if sheet_name not in sheet_items:
        raise KeyError(f"sheet not found: {sheet_name}")
    rows = sheet_items[sheet_name]
    starts: list[tuple[int, SectionLabel]] = []
    for index, row in enumerate(rows):
        label = _parse_section_label(row[0] if row else None)
        if label:
            starts.append((index, label))

    curves: dict[tuple[str, str, float, float], SpectrumCurve] = {}
    for offset, (start, label) in enumerate(starts):
        end = starts[offset + 1][0] if offset + 1 < len(starts) else len(rows)
        frequencies: list[float] = []
        values_by_damping: dict[float, list[float]] = {item: [] for item in DAMPING_COLUMNS}
        for row_number, row in enumerate(rows[start:end], start=start + 1):
            if not row or len(row) < 8:
                continue
            frequency = row[1]
            if not isinstance(frequency, int | float):
                continue
            if row[0] and 0.02 - 1e-12 <= float(frequency) <= 0.10 + 1e-12:
                continue
            frequencies.append(float(frequency))
            for damping_index, damping in enumerate(DAMPING_COLUMNS):
                value = row[2 + damping_index] if len(row) > 2 + damping_index else None
                values_by_damping[damping].append(float(value) if isinstance(value, int | float) else 0.0)
        if not frequencies:
            continue
        for damping, values in values_by_damping.items():
            if len(values) == len(frequencies):
                curves[(label.level, label.direction, label.elevation_m, damping)] = SpectrumCurve(
                    tuple(frequencies),
                    tuple(values),
                    f"{path.name}:{sheet_name}:section@row{start + 1}",
                )
    return curves


def _generic_preview_elevations(rows: list[tuple[Any, ...]]) -> list[float]:
    elevations: set[float] = set()
    marker = re.compile(r"(标高|楼层|生根|elevation|el\.?|m|米)", re.IGNORECASE)
    number = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)")
    for row in rows[:300]:
        for value in row[:80]:
            if value is None:
                continue
            text = str(value).strip()
            if not text or not marker.search(text):
                continue
            for match in number.finditer(text):
                try:
                    candidate = float(match.group(0))
                except ValueError:
                    continue
                if -100.0 <= candidate <= 300.0:
                    elevations.add(candidate)
    return sorted(elevations)


def _split_sheet_building_and_project(sheet_name: str) -> tuple[str, str]:
    if "_" not in sheet_name:
        return sheet_name, ""
    building, suffix = sheet_name.rsplit("_", 1)
    if re.fullmatch(r"[A-Za-z0-9]+", suffix or ""):
        return building, suffix
    return sheet_name, ""


def _generic_sheet_preview(sheet_name: str, rows: list[tuple[Any, ...]]) -> dict[str, Any] | None:
    elevations = _generic_preview_elevations(rows)
    if not elevations:
        return None
    building_guess, project_code_guess = _split_sheet_building_and_project(sheet_name)
    return {
        "sheet": sheet_name,
        "building_guess": building_guess,
        "project_code_guess": project_code_guess,
        "elevations": elevations,
        "levels": [],
        "directions": [],
        "section_count": 0,
        "preview_mode": "generic_cell_scan",
    }


def _all_segmented_sheets(path: Path) -> set[str]:
    names: set[str] = set()
    for sheet_name, rows in _read_workbook_rows(path):
        if _sheet_has_segmented_spectrum(rows):
            names.add(sheet_name)
    return names


def describe_segmented_spectrum_workbook(path: Path | str) -> dict[str, Any]:
    path = Path(path)
    sheets: list[dict[str, Any]] = []
    generic_sheets: list[dict[str, Any]] = []
    for sheet_name, rows in _read_workbook_rows(path):
        labels = [label for row in rows if (label := _parse_section_label(row[0] if row else None))]
        if not labels:
            generic = _generic_sheet_preview(sheet_name, rows)
            if generic:
                generic_sheets.append(generic)
            continue
        elevations = sorted({label.elevation_m for label in labels})
        levels = sorted({label.level for label in labels})
        directions = sorted({label.direction for label in labels})
        building_guess, project_code_guess = _split_sheet_building_and_project(sheet_name)
        sheets.append(
            {
                "sheet": sheet_name,
                "building_guess": building_guess,
                "project_code_guess": project_code_guess,
                "elevations": elevations,
                "levels": levels,
                "directions": directions,
                "section_count": len(labels),
                "preview_mode": "segmented_spectrum",
            }
        )
    preview_sheets = sheets or generic_sheets
    return {
        "status": "pass" if sheets else ("preview_only" if generic_sheets else "blocked"),
        "workbook": str(path),
        "workbook_sha256": sha256_file(path) if path.exists() else "",
        "sheet_count": len(preview_sheets),
        "sheets": preview_sheets,
        "available_buildings": sorted({item["building_guess"] for item in preview_sheets}),
        "available_elevations": sorted({elevation for item in preview_sheets for elevation in item.get("elevations", [])}),
        "policy": "Workbook preview discovers segmented spectrum sheets when possible. If the workbook uses another layout, generic preview lists likely sheet names and elevations only; real ANSYS remains blocked until the spectrum config is confirmed.",
    }


def _load_aliases(alias_path: Path = DEFAULT_ALIAS_PATH) -> list[dict[str, Any]]:
    if not alias_path.exists():
        return []
    payload = json.loads(alias_path.read_text(encoding="utf-8"))
    return [item for item in payload.get("aliases", []) if item.get("match") and item.get("sheet_template")]


def resolve_segmented_spectrum_sheet(
    workbook_path: Path | str,
    *,
    building: str,
    project_code: str,
    alias_path: Path | str = DEFAULT_ALIAS_PATH,
) -> str:
    workbook_path = Path(workbook_path)
    available = _all_segmented_sheets(workbook_path)
    if not available:
        raise ValueError(f"{workbook_path.name}: no segmented spectrum sheets found")

    raw = str(building or "").strip()
    direct_candidates = [
        raw,
        f"{raw}_{project_code}",
    ]
    for candidate in direct_candidates:
        if candidate in available:
            return candidate

    normalised = _normalise_text(raw)
    def alias_candidates(template: str) -> list[str]:
        formatted = template.format(project_code=project_code)
        candidates = [formatted]
        suffix = f"_{project_code}"
        if formatted.endswith(suffix):
            candidates.append(formatted[: -len(suffix)])
        elif project_code:
            candidates.append(f"{formatted}_{project_code}")
        return list(dict.fromkeys(candidate for candidate in candidates if candidate))

    for alias in sorted(_load_aliases(Path(alias_path)), key=lambda item: len(str(item["match"])), reverse=True):
        match = _normalise_text(alias["match"])
        if normalised == match or normalised.startswith(match):
            for candidate in alias_candidates(str(alias["sheet_template"])):
                if candidate in available:
                    return candidate

    for sheet_name in available:
        base = sheet_name.rsplit("_", 1)[0]
        if _normalise_text(base) and _normalise_text(base) in normalised:
            return sheet_name

    raise ValueError(f"No spectrum sheet matched building={building!r}, project_code={project_code!r}")


def _interp_linear(x0: float, y0: float, x1: float, y1: float, x: float) -> float:
    if abs(x1 - x0) < 1e-12:
        return y0
    return y0 + (y1 - y0) * ((x - x0) / (x1 - x0))


def _interp_log_frequency(curve: SpectrumCurve, target_frequency: float) -> float:
    frequencies = curve.frequency_hz
    values = curve.acceleration_g
    if target_frequency <= frequencies[0]:
        return values[0]
    if target_frequency >= frequencies[-1]:
        return values[-1]
    for index in range(len(frequencies) - 1):
        f0, f1 = frequencies[index], frequencies[index + 1]
        if f0 <= target_frequency <= f1:
            return 10 ** _interp_linear(
                math.log10(max(f0, 1e-12)),
                math.log10(max(values[index], 1e-12)),
                math.log10(max(f1, 1e-12)),
                math.log10(max(values[index + 1], 1e-12)),
                math.log10(max(target_frequency, 1e-12)),
            )
    return values[-1]


def _curve_at_elevation(
    curves: dict[tuple[str, str, float, float], SpectrumCurve],
    *,
    level: str,
    direction: str,
    elevation: float,
    damping: float,
) -> SpectrumCurve:
    candidates = sorted(
        (key[2], curve)
        for key, curve in curves.items()
        if key[0] == level and key[1] == direction and abs(key[3] - damping) < 1e-12
    )
    if not candidates:
        raise ValueError(f"No curve for level={level}, direction={direction}, damping={damping}")
    for candidate_elevation, curve in candidates:
        if abs(candidate_elevation - elevation) < 1e-8:
            return curve
    available = [item[0] for item in candidates]
    raise ValueError(
        f"No exact spectrum elevation for level={level}, direction={direction}, "
        f"damping={damping}, elevation={elevation:g}; available={available}"
    )


def resolve_spectrum_elevation(
    curves: dict[tuple[str, str, float, float], SpectrumCurve],
    *,
    elevation: float,
    requirements: list[tuple[str, str, float]] | tuple[tuple[str, str, float], ...],
) -> dict[str, Any]:
    """Resolve the workbook elevation actually used by spectrum extraction.

    Production policy: do not interpolate between floor spectra. Exact workbook
    elevations are used directly. Otherwise choose the lowest common workbook
    elevation that is not more than 0.1 m below the intake elevation. If the
    nearest lower floor is more than 0.1 m below the intake elevation, choose
    the next higher common floor.
    """

    requested = float(elevation)
    lower_tolerance = SPECTRUM_ELEVATION_LOWER_TOLERANCE_M
    minimum_allowed = requested - lower_tolerance
    requirement_payloads: list[dict[str, Any]] = []
    common_elevations: set[float] | None = None
    for level, direction, damping in requirements:
        available = sorted(
            key[2]
            for key in curves
            if key[0] == level and key[1] == direction and abs(key[3] - damping) < 1e-12
        )
        if not available:
            raise ValueError(f"No spectrum curves for level={level}, direction={direction}, damping={damping}")
        available_set = set(available)
        common_elevations = available_set if common_elevations is None else common_elevations.intersection(available_set)
        lower = [item for item in available if item < requested]
        upper = [item for item in available if item > requested]
        selectable = [item for item in available if item >= minimum_allowed - 1e-8]
        requirement_payloads.append(
            {
                "level": level,
                "direction": direction,
                "damping": damping,
                "available_elevations": available,
                "lower_elevation": max(lower) if lower else None,
                "upper_elevation": min(upper) if upper else None,
                "minimum_allowed_elevation": minimum_allowed,
                "selected_candidate_elevation": min(selectable) if selectable else None,
            }
        )

    common = sorted(common_elevations or [])
    selectable_common = [item for item in common if item >= minimum_allowed - 1e-8]
    if not selectable_common:
        raise ValueError(
            f"No common spectrum elevation >= requested elevation minus {lower_tolerance:g}m tolerance "
            f"for requested elevation={requested:g}; common_available={common}"
        )

    selected = float(selectable_common[0])
    if abs(selected - requested) < 1e-8:
        mode = "exact"
        policy_reason = "exact_workbook_elevation"
    elif selected < requested:
        mode = "lower_floor_within_0p1m_tolerance"
        policy_reason = "selected_lower_floor_is_within_0p1m_below_requested_elevation"
    else:
        mode = "next_upper_floor"
        policy_reason = "nearest_lower_floor_is_more_than_0p1m_below_requested_elevation_or_missing"
    lower_values = sorted({item["lower_elevation"] for item in requirement_payloads if item["lower_elevation"] is not None})
    upper_values = sorted({item["upper_elevation"] for item in requirement_payloads if item["upper_elevation"] is not None})
    payload: dict[str, Any] = {
        "mode": mode,
        "requested_elevation": requested,
        "selected_elevation": selected,
        "lower_tolerance_m": lower_tolerance,
        "minimum_allowed_elevation": minimum_allowed,
        "common_available_elevations": common,
        "requirements": requirement_payloads,
        "policy_reason": policy_reason,
    }
    if lower_values:
        payload["lower_elevations"] = lower_values
    if upper_values:
        payload["upper_elevations"] = upper_values
    return payload


def derive_static_acceleration_coefficients(
    workbook_path: Path | str,
    *,
    project_code: str,
    building: str,
    elevation: float,
    elevations: list[float] | tuple[float, ...] | None = None,
    alias_path: Path | str = DEFAULT_ALIAS_PATH,
    static_factor: float = 1.5,
    coefficient_source: str = "peak",
) -> dict[str, Any]:
    """Derive equivalent-static acceleration coefficients from a segmented spectrum workbook.

    Steel-platform static-method command streams use the peak acceleration of each selected
    spectrum curve. Response-spectrum static correction load steps use the high-frequency
    tail/zero-period acceleration used by the audited source command streams. Horizontal
    X/Y curves are enveloped and applied to both horizontal directions. The returned
    coefficients are unitless g values; the solver template applies
    `static_factor * coefficient * 9.81`.
    """
    if coefficient_source not in {"peak", "zero_period", "frequency_100hz"}:
        raise ValueError("coefficient_source must be 'peak', 'zero_period', or 'frequency_100hz'")

    workbook_path = Path(workbook_path)
    sheet = resolve_segmented_spectrum_sheet(
        workbook_path,
        building=building,
        project_code=project_code,
        alias_path=alias_path,
    )
    curves = _read_segmented_sheet(workbook_path, sheet)
    elevation_list = sorted({float(item) for item in (elevations or [elevation])})
    if not elevation_list:
        elevation_list = [float(elevation)]
    requirements = tuple(
        (level, direction, DAMPING_BY_LEVEL[level])
        for level in ("SL-1", "SL-2")
        for direction in ("X", "Y", "Z")
    )
    elevation_selections = [
        resolve_spectrum_elevation(curves, elevation=item, requirements=requirements)
        for item in elevation_list
    ]
    selected_elevation_list = sorted({float(item["selected_elevation"]) for item in elevation_selections})
    elevation_policy_suffix = (
        "with_explicit_elevation_envelope"
        if len(selected_elevation_list) > 1
        else "at_selected_spectrum_elevation"
    )
    payload: dict[str, Any] = {
        "status": "pass",
        "workbook": str(workbook_path),
        "workbook_sha256": sha256_file(workbook_path),
        "sheet": sheet,
        "requested_elevation": float(elevation),
        "elevation": float(selected_elevation_list[0]) if len(selected_elevation_list) == 1 else selected_elevation_list,
        "selected_elevation": float(selected_elevation_list[0]) if len(selected_elevation_list) == 1 else selected_elevation_list,
        "requested_elevations": elevation_list,
        "elevations": selected_elevation_list,
        "elevation_selections": elevation_selections,
        "static_acceleration_factor": static_factor,
        "coefficient_source": coefficient_source,
        "source_policy": (
            f"equivalent_static_uses_peak_spectral_acceleration_with_horizontal_envelope_{elevation_policy_suffix}"
            if coefficient_source == "peak"
            else (
                f"response_spectrum_static_correction_uses_100hz_acceleration_with_horizontal_envelope_{elevation_policy_suffix}"
                if coefficient_source == "frequency_100hz"
                else f"response_spectrum_static_correction_uses_high_frequency_zero_period_tail_with_horizontal_envelope_{elevation_policy_suffix}"
            )
        ),
        "source_refs": [],
    }
    for level, prefix in (("SL-1", "obe"), ("SL-2", "sse")):
        damping = DAMPING_BY_LEVEL[level]
        horizontal_candidates: list[float] = []
        vertical_candidates: list[float] = []
        for item_elevation in selected_elevation_list:
            x_curve = _curve_at_elevation(curves, level=level, direction="X", elevation=item_elevation, damping=damping)
            y_curve = _curve_at_elevation(curves, level=level, direction="Y", elevation=item_elevation, damping=damping)
            z_curve = _curve_at_elevation(curves, level=level, direction="Z", elevation=item_elevation, damping=damping)
            if coefficient_source == "frequency_100hz":
                horizontal_candidates.extend([x_curve.acceleration_at_hz(100.0), y_curve.acceleration_at_hz(100.0)])
                vertical_candidates.append(z_curve.acceleration_at_hz(100.0))
            elif coefficient_source == "zero_period":
                horizontal_candidates.extend([x_curve.zero_period_g, y_curve.zero_period_g])
                vertical_candidates.append(z_curve.zero_period_g)
            else:
                horizontal_candidates.extend([x_curve.peak_g, y_curve.peak_g])
                vertical_candidates.append(z_curve.peak_g)
            payload["source_refs"].extend([x_curve.source_ref, y_curve.source_ref, z_curve.source_ref])
        horizontal = max(horizontal_candidates)
        vertical = max(vertical_candidates)
        payload[f"zpa_{prefix}_x_g"] = horizontal
        payload[f"zpa_{prefix}_y_g"] = horizontal
        payload[f"zpa_{prefix}_z_g"] = vertical
    return payload
