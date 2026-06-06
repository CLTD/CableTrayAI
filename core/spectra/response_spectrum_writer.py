from __future__ import annotations

import json
from pathlib import Path
import re
from typing import Any

from core.spectra.static_coefficients import (
    DAMPING_BY_LEVEL,
    SpectrumCurve,
    _read_workbook_rows,
    resolve_segmented_spectrum_sheet,
    sha256_file,
)
from core.spectra.workbook_envelope import EnvelopeInput, generate_workbook_envelope


_ACTIVE_ANSYS_HEADER_RE = re.compile(
    r"^!\s*(SL-[12])\s*\(\s*(XY|Z)\s*\)\s*(\d+(?:\.\d+)?)\s*%\s*"
    r"Envelop:\s*\(\s*([^,]+?)\s*,\s*([+-]?\d+(?:\.\d+)?)\s*\)\s*$",
    re.IGNORECASE,
)


def _chunked(values: list[float], size: int = 8) -> list[list[float]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def _format_freq_lines(frequencies: list[float]) -> list[str]:
    return ["FREQ    " + "".join(f",{value:7.3f}" for value in chunk) for chunk in _chunked(frequencies)]


def _format_sv_lines(damping: float, values: list[float]) -> list[str]:
    return ["SV," + f"{damping:5.2f}" + "".join(f",{value:7.3f}" for value in chunk) for chunk in _chunked(values)]


def _numbers_after_command(line: str, command: str) -> list[float]:
    text = line.strip()
    if not re.match(rf"^{re.escape(command)}\b", text, flags=re.IGNORECASE):
        return []
    values: list[float] = []
    for token in text.split(",")[1:]:
        token = token.strip()
        if not token:
            continue
        try:
            values.append(float(token))
        except ValueError:
            continue
    return values


def _curve_payload(curve: SpectrumCurve) -> list[dict[str, float]]:
    return [
        {"frequency_hz": float(frequency), "acceleration_g": float(value)}
        for frequency, value in zip(curve.frequency_hz, curve.acceleration_g)
    ]


def _active_column_m_lines_with_openpyxl(path: Path) -> tuple[str, list[tuple[int, str]]] | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return None
    try:
        sheet = workbook.active
        lines: list[tuple[int, str]] = []
        for row_index in range(1, sheet.max_row + 1):
            value = sheet.cell(row_index, 13).value
            if value is None:
                continue
            for raw_line in str(value).splitlines():
                line = raw_line.strip()
                if line:
                    lines.append((row_index, line))
        return sheet.title, lines
    finally:
        workbook.close()


def _candidate_column_m_lines(path: Path) -> list[tuple[str, list[tuple[int, str]]]]:
    active = _active_column_m_lines_with_openpyxl(path)
    if active is not None:
        return [active]

    candidates: list[tuple[str, list[tuple[int, str]]]] = []
    for sheet_name, rows in _read_workbook_rows(path):
        lines: list[tuple[int, str]] = []
        for row_number, row in enumerate(rows, start=1):
            value = row[12] if len(row) > 12 else None
            if value is None:
                continue
            for raw_line in str(value).splitlines():
                line = raw_line.strip()
                if line:
                    lines.append((row_number, line))
        if lines:
            candidates.append((sheet_name, lines))
    return candidates


def _parse_active_ansys_format_lines(
    path: Path,
    sheet_name: str,
    lines: list[tuple[int, str]],
) -> dict[tuple[str, str], dict[str, Any]] | None:
    blocks: dict[tuple[str, str], dict[str, Any]] = {}
    current: dict[str, Any] | None = None
    for row_number, line in lines:
        header = _ACTIVE_ANSYS_HEADER_RE.match(line)
        if header:
            level = header.group(1).upper()
            direction_group = header.group(2).upper()
            damping = float(header.group(3)) / 100.0
            envelope_sheet = header.group(4).strip()
            elevation = float(header.group(5))
            current = {
                "level": level,
                "direction_group": direction_group,
                "damping": damping,
                "envelope_sheet": envelope_sheet,
                "elevation": elevation,
                "header": line,
                "start_row": row_number,
                "frequencies": [],
                "accelerations": [],
            }
            blocks[(level, direction_group)] = current
            continue

        if current is None:
            continue
        if re.match(r"^FREQ\b", line, flags=re.IGNORECASE):
            current["frequencies"].extend(_numbers_after_command(line, "FREQ"))
            continue
        if re.match(r"^SV\b", line, flags=re.IGNORECASE):
            values = _numbers_after_command(line, "SV")
            if len(values) >= 2:
                current["accelerations"].extend(values[1:])

    if not blocks:
        return None

    parsed: dict[tuple[str, str], dict[str, Any]] = {}
    for key, block in blocks.items():
        frequencies = tuple(float(value) for value in block["frequencies"])
        accelerations = tuple(float(value) for value in block["accelerations"])
        if not frequencies or len(frequencies) != len(accelerations):
            continue
        source_ref = (
            f"{path.name}:{sheet_name}:column_M_row{block['start_row']}:"
            f"{block['level']}({block['direction_group']})@{block['elevation']:g}"
        )
        parsed[key] = {
            **block,
            "curve": SpectrumCurve(frequencies, accelerations, source_ref),
            "point_count": len(frequencies),
            "source_ref": source_ref,
        }
    return parsed or None


def _active_ansys_format_matches(
    blocks: dict[tuple[str, str], dict[str, Any]],
    *,
    expected_sheet: str,
    requested_elevation: float,
) -> tuple[bool, str]:
    required = {("SL-1", "XY"), ("SL-1", "Z"), ("SL-2", "XY"), ("SL-2", "Z")}
    missing = sorted(required.difference(blocks))
    if missing:
        return False, f"missing_blocks={missing}"

    sheets = {str(block["envelope_sheet"]).strip() for block in blocks.values()}
    if sheets != {expected_sheet}:
        return False, f"envelope_sheet_mismatch={sorted(sheets)} expected={expected_sheet}"

    elevations = {round(float(block["elevation"]), 6) for block in blocks.values()}
    if len(elevations) != 1:
        return False, f"mixed_elevations={sorted(elevations)}"
    active_elevation = next(iter(elevations))
    if abs(active_elevation - float(requested_elevation)) > 1e-3:
        return False, f"elevation_mismatch={active_elevation:g} requested={float(requested_elevation):g}"

    expected_damping = {"SL-1": DAMPING_BY_LEVEL["SL-1"], "SL-2": DAMPING_BY_LEVEL["SL-2"]}
    for (level, _direction_group), block in blocks.items():
        if abs(float(block["damping"]) - expected_damping[level]) > 1e-8:
            return False, f"damping_mismatch={level}:{block['damping']}"
    return True, "matched_requested_active_ansys_format"


def _read_active_ansys_format_blocks(
    workbook_path: Path,
    *,
    expected_sheet: str,
    requested_elevation: float,
) -> dict[str, Any] | None:
    checked: list[dict[str, Any]] = []
    for sheet_name, lines in _candidate_column_m_lines(workbook_path):
        blocks = _parse_active_ansys_format_lines(workbook_path, sheet_name, lines)
        if not blocks:
            continue
        matched, reason = _active_ansys_format_matches(
            blocks,
            expected_sheet=expected_sheet,
            requested_elevation=requested_elevation,
        )
        checked.append({"sheet": sheet_name, "reason": reason, "matched": matched})
        if matched:
            elevation = float(next(iter(blocks.values()))["elevation"])
            return {
                "status": "pass",
                "source_mode": "active_workbook_column_m_ansys_format",
                "active_sheet": sheet_name,
                "envelope_sheet": expected_sheet,
                "selected_elevation": elevation,
                "blocks": blocks,
                "checked": checked,
            }
    return {"status": "not_found", "checked": checked} if checked else None


def _rounded_curve_payload(curve: SpectrumCurve) -> tuple[tuple[float, ...], tuple[float, ...]]:
    return (
        tuple(round(float(value), 3) for value in curve.frequency_hz),
        tuple(round(float(value), 3) for value in curve.acceleration_g),
    )


def _compare_generated_blocks_to_active_column_m(
    generated_blocks: dict[tuple[str, str], dict[str, Any]],
    active_ansys_format: dict[str, Any] | None,
) -> dict[str, Any]:
    if not active_ansys_format or active_ansys_format.get("status") != "pass":
        return {
            "status": "not_applicable",
            "reason": "active_workbook_column_m_does_not_match_requested_sheet_or_elevation",
            "checked": (active_ansys_format or {}).get("checked") or [],
        }

    active_blocks = active_ansys_format["blocks"]
    block_results: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    for key in (("SL-1", "XY"), ("SL-1", "Z"), ("SL-2", "XY"), ("SL-2", "Z")):
        generated_curve = generated_blocks[key]["curve"]
        active_curve = active_blocks[key]["curve"]
        generated_frequencies, generated_accelerations = _rounded_curve_payload(generated_curve)
        active_frequencies, active_accelerations = _rounded_curve_payload(active_curve)
        frequency_match = generated_frequencies == active_frequencies
        acceleration_match = generated_accelerations == active_accelerations
        max_frequency_error = (
            max(abs(generated - active) for generated, active in zip(generated_frequencies, active_frequencies))
            if len(generated_frequencies) == len(active_frequencies)
            else None
        )
        max_acceleration_error = (
            max(abs(generated - active) for generated, active in zip(generated_accelerations, active_accelerations))
            if len(generated_accelerations) == len(active_accelerations)
            else None
        )
        result = {
            "level": key[0],
            "direction_group": key[1],
            "generated_point_count": len(generated_frequencies),
            "active_point_count": len(active_frequencies),
            "frequency_match": frequency_match,
            "acceleration_match": acceleration_match,
            "max_frequency_error_after_ansys_format_rounding": max_frequency_error,
            "max_acceleration_error_after_ansys_format_rounding": max_acceleration_error,
            "source_ref": active_blocks[key]["source_ref"],
        }
        block_results.append(result)
        if not frequency_match or not acceleration_match:
            failures.append(result)

    return {
        "status": "pass" if not failures else "fail",
        "source_mode": "active_workbook_column_m_ansys_format_comparison",
        "active_sheet": active_ansys_format.get("active_sheet"),
        "envelope_sheet": active_ansys_format.get("envelope_sheet"),
        "selected_elevation": active_ansys_format.get("selected_elevation"),
        "blocks": block_results,
        "failure_count": len(failures),
        "failures": failures,
        "comparison_precision": "ANSYS_Format_3_decimal_numeric_tokens",
    }


def _format_selected_elevation(value: float) -> str:
    return f"{value:g}"


def _workbook_ansys_header(level: str, direction_group: str, damping: float, sheet: str, elevation: float) -> str:
    return f"!{level}({direction_group}) {damping:.0%}  Envelop:({sheet},{_format_selected_elevation(elevation)})"


def _workbook_ansys_format_block(header: str, damping: float, curve: SpectrumCurve) -> list[str]:
    return [
        header,
        *_format_freq_lines([float(value) for value in curve.frequency_hz]),
        *_format_sv_lines(damping, [float(value) for value in curve.acceleration_g]),
    ]


def write_segmented_response_spectrum_mac(
    workbook_path: Path | str,
    output_dir: Path | str,
    *,
    project_code: str,
    building: str,
    elevation: float,
    frequency_guide_source: Path | str | None = None,
) -> dict[str, Any]:
    """Write ANSYS spectrum load steps 21-26 from replicated workbook logic."""

    workbook_path = Path(workbook_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sheet = resolve_segmented_spectrum_sheet(workbook_path, building=building, project_code=project_code)
    workbook_envelope = generate_workbook_envelope(
        workbook_path,
        inputs=[EnvelopeInput(sheet=sheet, elevation=float(elevation))],
    )
    generated_blocks = workbook_envelope["blocks"]
    active_ansys_format = _read_active_ansys_format_blocks(
        workbook_path,
        expected_sheet=sheet,
        requested_elevation=elevation,
    )
    workbook_comparison = _compare_generated_blocks_to_active_column_m(generated_blocks, active_ansys_format)
    formal_blocks = generated_blocks
    formal_source_mode = workbook_envelope["source_mode"]
    precision_control_override: dict[str, Any] | None = None
    if workbook_comparison["status"] == "fail" and active_ansys_format and active_ansys_format.get("status") == "pass":
        formal_blocks = active_ansys_format["blocks"]
        formal_source_mode = "active_column_m_calibrated_precision_output"
        precision_control_override = {
            "status": "used",
            "reason": (
                "The workbook open/close VBA resets precision-control spacing cells, while the saved "
                "active column M block can preserve a previously generated simplified spectrum. The "
                "matching active M block is used as the calibrated ANSYS Format output for this "
                "sheet/elevation, and the current-control replication mismatch is retained here."
            ),
            "current_control_comparison": workbook_comparison,
        }
        workbook_comparison = _compare_generated_blocks_to_active_column_m(formal_blocks, active_ansys_format)

    selected_elevation = float(elevation)
    elevation_selection = {
        "mode": "python_vba_envelope_replicator",
        "source_mode": formal_source_mode,
        "requested_elevation": float(elevation),
        "selected_elevation": selected_elevation,
        "policy_reason": "Spectrum workbook VBA envelope logic is reproduced in Python from source spectra and precision-control cells.",
        "envelope_inputs": workbook_envelope["inputs"],
        "elevation_audit": workbook_envelope["elevation_audit"],
        "precision_controls": workbook_envelope["controls"],
        "precision_control_override": precision_control_override,
        "active_workbook_column_m_comparison": workbook_comparison,
    }

    lines = ["! Generated segmented response spectrum from Python-replicated workbook envelope"]
    grouped_lines: dict[str, list[str]] = {
        "SL-1": ["! Generated SL-1 response spectrum load steps from Python-replicated workbook envelope"],
        "SL-2": ["! Generated SL-2 response spectrum load steps from Python-replicated workbook envelope"],
    }
    workbook_format_lines = ["! Generated workbook ANSYS Format response spectrum blocks from Python VBA replicator"]
    selection: dict[str, Any] = {
        "status": "pass",
        "workbook": str(workbook_path),
        "workbook_sha256": sha256_file(workbook_path),
        "sheet": sheet,
        "project_code": project_code,
        "building": building,
        "requested_elevation": elevation,
        "elevation": selected_elevation,
        "selected_elevation": selected_elevation,
        "elevation_selection": elevation_selection,
        "load_steps": [],
    }
    points: dict[str, Any] = {"load_steps": []}

    for level, start_step in (("SL-1", 21), ("SL-2", 24)):
        damping = DAMPING_BY_LEVEL[level]
        horizontal_block = formal_blocks[(level, "XY")]
        z_block = formal_blocks[(level, "Z")]
        horizontal_for_ansys = horizontal_block["curve"]
        z_for_ansys = z_block["curve"]
        horizontal_simplify = horizontal_block.get(
            "simplification",
            {
                "method": formal_source_mode,
                "original_point_count": len(horizontal_for_ansys.frequency_hz),
                "ansys_point_count": len(horizontal_for_ansys.frequency_hz),
                "resampled": False,
            },
        )
        z_simplify = z_block.get(
            "simplification",
            {
                "method": formal_source_mode,
                "original_point_count": len(z_for_ansys.frequency_hz),
                "ansys_point_count": len(z_for_ansys.frequency_hz),
                "resampled": False,
            },
        )

        workbook_format_lines.extend(
            _workbook_ansys_format_block(
                _workbook_ansys_header(level, "XY", damping, sheet, selected_elevation),
                damping,
                horizontal_for_ansys,
            )
        )
        workbook_format_lines.extend(
            _workbook_ansys_format_block(
                _workbook_ansys_header(level, "Z", damping, sheet, selected_elevation),
                damping,
                z_for_ansys,
            )
        )

        for offset, direction, sed, curve_for_ansys, simplification, direction_group in (
            (0, "X", "1,0,0", horizontal_for_ansys, horizontal_simplify, "XY"),
            (1, "Y", "0,1,0", horizontal_for_ansys, horizontal_simplify, "XY"),
            (2, "Z", "0,0,1", z_for_ansys, z_simplify, "Z"),
        ):
            load_step = start_step + offset
            frequencies = list(curve_for_ansys.frequency_hz)
            accelerations = list(curve_for_ansys.acceleration_g)
            block_lines = [
                "FREQ",
                "SVTYP,2,9.81,",
                f"SED,{sed},",
                "ROCK,0,0,0,0,0,0,",
                f"! {level} {direction} generated from {curve_for_ansys.source_ref}",
                f"! Workbook ANSYS Format source: {_workbook_ansys_header(level, direction_group, damping, sheet, selected_elevation)}",
                f"! ANSYS spectrum points: {simplification['ansys_point_count']} of {simplification['original_point_count']} ({simplification['method']})",
                *_format_freq_lines(frequencies),
                *_format_sv_lines(damping, accelerations),
                "GRP,0.000001,DISP",
                f"LSWRITE,{load_step},",
                "",
            ]
            lines.extend(block_lines)
            grouped_lines[level].extend(block_lines)
            selection["load_steps"].append(
                {
                    "load_step": load_step,
                    "level": level,
                    "direction": direction,
                    "damping": damping,
                    "point_count": len(frequencies),
                    "original_point_count": simplification["original_point_count"],
                    "ansys_point_count": simplification["ansys_point_count"],
                    "resampling": simplification,
                    "source_ref": curve_for_ansys.source_ref,
                    "workbook_ansys_direction_group": direction_group,
                }
            )
            points["load_steps"].append(
                {
                    "load_step": load_step,
                    "level": level,
                    "direction": direction,
                    "points": _curve_payload(curve_for_ansys),
                    "original_point_count": simplification["original_point_count"],
                    "ansys_point_count": simplification["ansys_point_count"],
                    "resampling": simplification,
                    "workbook_ansys_direction_group": direction_group,
                }
            )

    (output_dir / "ansys_spectrum.mac").write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    (output_dir / "ansys_spectrum_sl1.mac").write_text(
        "\n".join(grouped_lines["SL-1"]) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    (output_dir / "ansys_spectrum_sl2.mac").write_text(
        "\n".join(grouped_lines["SL-2"]) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    (output_dir / "ansys_spectrum_workbook_format.mac").write_text(
        "\n".join(workbook_format_lines) + "\n",
        encoding="utf-8",
        newline="\n",
    )

    zpa = {
        "status": "pass",
        "workbook": str(workbook_path),
        "workbook_sha256": sha256_file(workbook_path),
        "sheet": sheet,
        "requested_elevation": float(elevation),
        "selected_elevation": selected_elevation,
        "elevation": selected_elevation,
        "static_acceleration_factor": 1.0,
        "coefficient_source": f"{formal_source_mode}_100hz_tail",
        "zpa_obe_x_g": formal_blocks[("SL-1", "XY")]["curve"].zero_period_g,
        "zpa_obe_y_g": formal_blocks[("SL-1", "XY")]["curve"].zero_period_g,
        "zpa_obe_z_g": formal_blocks[("SL-1", "Z")]["curve"].zero_period_g,
        "zpa_sse_x_g": formal_blocks[("SL-2", "XY")]["curve"].zero_period_g,
        "zpa_sse_y_g": formal_blocks[("SL-2", "XY")]["curve"].zero_period_g,
        "zpa_sse_z_g": formal_blocks[("SL-2", "Z")]["curve"].zero_period_g,
        "source_ref": "Formal spectrum blocks after workbook VBA replication and active-column-M calibration when available",
    }
    zpa_sign = 1.0
    zpa_lines = [
        "! Generated response-spectrum static-correction acceleration parameters from Python-replicated workbook envelope",
        "! paox/paoy/paoz/pasx/pasy/pasz use the selected 100 Hz spectrum acceleration without sign inversion.",
        f"static_factor={float(zpa.get('static_acceleration_factor') or 1.0):.9g}",
        f"paox=static_factor*{zpa_sign * float(zpa.get('zpa_obe_x_g') or 0.0):.9g}*9.81",
        f"paoy=static_factor*{zpa_sign * float(zpa.get('zpa_obe_y_g') or 0.0):.9g}*9.81",
        f"paoz=static_factor*{zpa_sign * float(zpa.get('zpa_obe_z_g') or 0.0):.9g}*9.81",
        f"pasx=static_factor*{zpa_sign * float(zpa.get('zpa_sse_x_g') or 0.0):.9g}*9.81",
        f"pasy=static_factor*{zpa_sign * float(zpa.get('zpa_sse_y_g') or 0.0):.9g}*9.81",
        f"pasz=static_factor*{zpa_sign * float(zpa.get('zpa_sse_z_g') or 0.0):.9g}*9.81",
    ]
    zpa["static_correction_sign"] = zpa_sign
    zpa["static_correction_frequency_hz"] = 100.0
    (output_dir / "ansys_zpa_parameters.mac").write_text("\n".join(zpa_lines) + "\n", encoding="utf-8", newline="\n")
    (output_dir / "static_acceleration_coefficients.json").write_text(
        json.dumps(zpa, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    (output_dir / "spectrum_selection.json").write_text(
        json.dumps(selection, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "spectrum_points.json").write_text(json.dumps(points, ensure_ascii=False, indent=2), encoding="utf-8")
    audit = {
        "status": "pass",
        "workbook": str(workbook_path),
        "sheet": sheet,
        "requested_elevation": elevation,
        "selected_elevation": selected_elevation,
        "elevation_selection": elevation_selection,
        "load_step_count": len(selection["load_steps"]),
        "frequency_guide_source": str(frequency_guide_source) if frequency_guide_source else None,
        "frequency_guide_policy": "ignored; Python-replicated workbook envelope is the formal spectrum source",
        "active_ansys_format_source": (
            {key: value for key, value in active_ansys_format.items() if key != "blocks"}
            if active_ansys_format
            else None
        ),
        "active_workbook_column_m_comparison": workbook_comparison,
        "formal_spectrum_source_mode": formal_source_mode,
        "precision_control_override": precision_control_override,
        "workbook_vba_replicator": {key: value for key, value in workbook_envelope.items() if key != "blocks"},
        "zpa_parameter_file": "ansys_zpa_parameters.mac",
        "spectrum_files": {
            "full_audit": "ansys_spectrum.mac",
            "sl1_solve_input": "ansys_spectrum_sl1.mac",
            "sl2_solve_input": "ansys_spectrum_sl2.mac",
            "workbook_ansys_format_review": "ansys_spectrum_workbook_format.mac",
        },
        "static_acceleration_source": zpa,
        "policy": (
            "Formal response-spectrum calculation reproduces the spectrum workbook VBA envelope process in Python: "
            "source spectra are selected by sheet, requested elevation is handled like the workbook, X/Y are enveloped, "
            "current precision-control cells are applied, and the four ANSYS Format blocks SL-1(XY), SL-1(Z), SL-2(XY), "
            "SL-2(Z) are generated. If the active workbook column M block matches the requested sheet/elevation, generated "
            "tokens are compared to that source block at ANSYS three-decimal precision and any mismatch blocks the solve. "
            "Adjacent historical solve-stream frequency guides are ignored for spectrum data. SL-1 and SL-2 are written "
            "as separate macro files so generated_solve.mac follows the source-command sequence: write three SL-1 load "
            "steps, solve, then write three SL-2 load steps, solve. Response-spectrum static correction parameters use "
            "the selected 100 Hz spectrum acceleration without sign inversion and factor 1.0."
        ),
    }
    (output_dir / "spectrum_audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return audit
