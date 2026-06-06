from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from core.intake.intake_excel_reader import _openpyxl_unavailable_due_to_expat, _read_xlsx_rows_without_expat
from core.spectra.config_schema import SpectrumFormatConfig, column_index


def sha256_file(path: Path | str) -> str:
    path = Path(path)
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _norm(value: Any) -> str:
    return str(value or "").strip().lower()


HEADER_ALIASES = {
    "project_code": {"project_code", "project", "项目"},
    "building": {"building", "厂房"},
    "area": {"area", "区域"},
    "elevation": {"elevation", "标高"},
    "damping": {"damping", "damping_ratio", "阻尼", "阻尼比"},
    "level": {"level", "spectrum_level", "sl"},
    "direction": {"direction", "方向"},
    "frequency_hz": {"frequency_hz", "freq", "frequency", "频率"},
    "acceleration_g": {"acceleration_g", "accel", "acceleration", "加速度"},
}


def _header_key(value: Any) -> str | None:
    normalized = _norm(value).replace(" ", "_")
    for key, aliases in HEADER_ALIASES.items():
        if normalized in {alias.lower() for alias in aliases}:
            return key
    return None


def _get_config_value(row: tuple, column: str | None, default: Any = None) -> Any:
    if not column:
        return default
    idx = column_index(column) - 1
    return row[idx] if idx < len(row) else default


def _normalise_rows(rows: list[Any]) -> list[tuple[Any, ...]]:
    return [tuple(row or ()) for row in rows]


def _read_workbook_rows_with_openpyxl(path: Path) -> list[tuple[str, list[tuple[Any, ...]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        return [
            (sheet.title, _normalise_rows(list(sheet.iter_rows(values_only=True))))
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _read_workbook_rows(path: Path) -> list[tuple[str, list[tuple[Any, ...]]]]:
    try:
        return _read_workbook_rows_with_openpyxl(path)
    except Exception as exc:
        if not _openpyxl_unavailable_due_to_expat(exc):
            raise
        return [(sheet_name, _normalise_rows(rows)) for sheet_name, rows in _read_xlsx_rows_without_expat(path)]


def _read_configured_table(path: Path, sheet_items: list[tuple[str, list[tuple[Any, ...]]]], config: SpectrumFormatConfig) -> dict | None:
    sheet_names = [name for name, _rows in sheet_items]
    sheet_name = config.sheet or config.sheet_pattern or sheet_names[0]
    if sheet_name not in sheet_names:
        return None
    rows = dict(sheet_items)[sheet_name]
    min_row = (config.header_row or 1) + 1
    points: list[dict] = []
    columns = config.columns
    defaults = config.defaults
    for row_index, row in enumerate(rows[min_row - 1 :], start=min_row):
        frequency = _get_config_value(row, columns.frequency_hz)
        if frequency in (None, ""):
            continue
        if columns.acceleration_columns:
            for direction, accel_col in columns.acceleration_columns.items():
                accel = _get_config_value(row, accel_col)
                if accel in (None, ""):
                    continue
                points.append(
                    {
                        "project_code": str(_get_config_value(row, columns.project_code, defaults.get("project_code")) or ""),
                        "building": str(_get_config_value(row, columns.building, defaults.get("building")) or ""),
                        "area": str(_get_config_value(row, columns.area, defaults.get("area")) or ""),
                        "elevation": float(_get_config_value(row, columns.elevation, defaults.get("elevation"))),
                        "damping": float(_get_config_value(row, columns.damping, defaults.get("damping"))),
                        "level": str(_get_config_value(row, columns.level, defaults.get("level"))),
                        "direction": str(direction).upper(),
                        "frequency_hz": float(frequency),
                        "acceleration_g": float(accel),
                        "source_ref": f"{path.name}:{sheet_name}:row{row_index}",
                    }
                )
        else:
            accel = _get_config_value(row, columns.acceleration_g)
            if accel in (None, ""):
                continue
            points.append(
                {
                    "project_code": str(_get_config_value(row, columns.project_code, defaults.get("project_code")) or ""),
                    "building": str(_get_config_value(row, columns.building, defaults.get("building")) or ""),
                    "area": str(_get_config_value(row, columns.area, defaults.get("area")) or ""),
                    "elevation": float(_get_config_value(row, columns.elevation, defaults.get("elevation"))),
                    "damping": float(_get_config_value(row, columns.damping, defaults.get("damping"))),
                    "level": str(_get_config_value(row, columns.level, defaults.get("level"))),
                    "direction": str(_get_config_value(row, columns.direction, defaults.get("direction"))).upper(),
                    "frequency_hz": float(frequency),
                    "acceleration_g": float(accel),
                    "source_ref": f"{path.name}:{sheet_name}:row{row_index}",
                }
            )
    if not points:
        raise ValueError(f"{path.name}: configured spectrum table produced no points")
    return {"workbook": str(path), "sheet": sheet_name, "sha256": sha256_file(path), "points": points}


def read_spectrum_workbook(path: Path | str, config: dict | SpectrumFormatConfig | None = None) -> dict:
    path = Path(path)
    sheet_items = _read_workbook_rows(path)
    if isinstance(config, dict):
        if "columns" in config:
            if hasattr(SpectrumFormatConfig, "model_validate"):
                config = SpectrumFormatConfig.model_validate(config)
            else:
                config = SpectrumFormatConfig.parse_obj(config)
        else:
            config = None
    if isinstance(config, SpectrumFormatConfig):
        configured = _read_configured_table(path, sheet_items, config)
        if configured is not None:
            return configured
    sheet_name, rows = sheet_items[0]

    header_row = None
    headers: list[str] = []
    for row_index, row in enumerate(rows, start=1):
        keys = [_header_key(value) for value in row]
        if sum(1 for key in keys if key is not None) >= 5:
            header_row = row_index
            headers = [key or f"unused_{index}" for index, key in enumerate(keys)]
            break
    if header_row is None:
        raise ValueError(f"{path.name}: no configurable spectrum table header found")

    points: list[dict] = []
    for row in rows[header_row:]:
        record = dict(zip(headers, row))
        if record.get("frequency_hz") in (None, "") or record.get("acceleration_g") in (None, ""):
            continue
        points.append(
            {
                "project_code": str(record.get("project_code") or ""),
                "building": str(record.get("building") or ""),
                "area": str(record.get("area") or ""),
                "elevation": float(record["elevation"]),
                "damping": float(record["damping"]),
                "level": str(record["level"]),
                "direction": str(record["direction"]).upper(),
                "frequency_hz": float(record["frequency_hz"]),
                "acceleration_g": float(record["acceleration_g"]),
                "source_ref": f"{path.name}:{sheet_name}:row{header_row + len(points) + 1}",
            }
        )
    if not points:
        raise ValueError(f"{path.name}: spectrum table header found but no points parsed")
    return {"workbook": str(path), "sheet": sheet_name, "sha256": sha256_file(path), "points": points}
