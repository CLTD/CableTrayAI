from __future__ import annotations

import csv
import html
import json
import re
import zipfile
from pathlib import Path
from typing import Any


REQUIRED_FIELDS = ["project_code", "building", "area", "elevation", "damping_ratio", "material"]

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "serial": ("\u5e8f\u53f7", "\u884c\u53f7", "\u9879\u53f7", "\u6761\u76ee", "serial", "row"),
    "report_number": (
        "\u8ba1\u7b97\u6279\u6b21",
        "\u62a5\u544a\u53f7",
        "\u62a5\u544a\u7f16\u53f7",
        "\u529b\u5b66\u62a5\u544a\u53f7",
        "\u529b\u5b66\u8ba1\u7b97\u7ed3\u679c",
        "\u8ba1\u7b97\u4e66\u53f7",
        "\u63d0\u8d44\u5355\u53f7",
        "\u7efc\u5408\u7f16\u53f7",
        "\u7efc\u5408\u53f7",
        "\u8ba1\u7b97\u7f16\u53f7",
        "\u8ba1\u7b97\u5355\u53f7",
        "\u63d0\u8d44\u7f16\u53f7",
        "\u63d0\u8d44\u6279\u6b21",
        "\u5355\u53f7",
        "report_number",
        "calculation_batch",
    ),
    "support_type": (
        "\u652f\u67b6\u5f62\u5f0f",
        "\u652f\u67b6\u7c7b\u578b",
        "\u652f\u67b6\u578b\u53f7",
        "\u652f\u540a\u67b6\u5f62\u5f0f",
        "\u652f\u540a\u67b6\u7c7b\u578b",
        "\u652f\u540a\u67b6\u578b\u53f7",
        "\u7c7b\u578b",
        "\u652f\u67b6",
        "support_type",
    ),
    "support_spacing_m": ("\u652f\u67b6\u95f4\u8ddd", "\u95f4\u8ddd", "\u8de8\u8ddd", "support_spacing_m"),
    "support_height_m": ("\u65b9\u94a2\u957f\u5ea6", "\u652f\u67b6\u9ad8\u5ea6", "\u9ad8\u5ea6", "\u957f\u5ea6", "support_height_m"),
    "description": (
        "\u6258\u76d8\u8f7d\u8377",
        "\u6258\u76d8\u8377\u8f7d",
        "\u6258\u76d8\u81ea\u91cd",
        "\u7535\u7f06\u6258\u76d8\u81ea\u91cd",
        "\u6258\u76d8\u7b49\u6548\u5bc6\u5ea6",
        "\u8f7d\u8377\u63cf\u8ff0",
        "\u8377\u8f7d\u63cf\u8ff0",
        "\u8f7d\u8377",
        "\u8377\u8f7d",
        "\u8bf4\u660e",
        "\u5907\u6ce8",
        "description",
    ),
    "square_section_spec": (
        "\u57cb\u4ef6",
        "\u57cb\u677f",
        "\u65b9\u94a2\u622a\u9762",
        "\u65b9\u94a2\u5c3a\u5bf8",
        "\u65b9\u94a2\u89c4\u683c",
        "\u65b9\u94a2\u7ba1\u89c4\u683c",
        "\u5efa\u8bae\u8ba1\u7b97\u65b9\u94a2",
        "\u6297\u9707\u8ba1\u7b97\u540e\u65b9\u94a2\u5c3a\u5bf8",
        "square_section_spec",
    ),
    "building": ("\u5382\u623f", "\u5382\u623f/\u8c31\u8868", "\u5382\u623f\u8c31\u8868", "\u8c31\u8868", "building"),
    "area": ("\u533a\u57df", "\u533a\u57df/\u8c31\u8868", "\u8c31\u8868\u533a\u57df", "area"),
    "elevation": ("\u751f\u6839\u5c42", "\u751f\u6839\u697c\u5c42", "\u6807\u9ad8", "\u5c42\u4f4d", "\u6807\u9ad8/\u697c\u5c42", "\u697c\u5c42", "elevation"),
    "support_id": ("\u652f\u67b6\u53f7", "\u652f\u540a\u67b6\u53f7", "\u540d\u79f0", "support_id"),
    "project_code": ("\u9879\u76ee\u53f7", "\u9879\u76ee\u4ee3\u7801", "project_code"),
    "damping_ratio": ("\u963b\u5c3c\u6bd4", "\u963b\u5c3c", "damping_ratio"),
    "material": ("\u6750\u6599", "\u6750\u8d28", "material"),
    "analysis_method": ("\u8ba1\u7b97\u65b9\u6cd5", "\u5206\u6790\u65b9\u6cd5", "analysis_method"),
    "cantilever_evaluation_mode": (
        "\u6258\u81c2\u8bc4\u5b9a\u65b9\u5f0f",
        "\u6258\u81c2\u5e94\u529b\u56fe",
        "\u6258\u81c2\u4e91\u56fe",
        "\u710a\u7f1d\u8bc4\u5b9a\u65b9\u5f0f",
        "cantilever_evaluation_mode",
    ),
}

STATIC_METHOD_KEYWORD = "\u94a2\u5e73\u53f0"


def _coerce(value: Any) -> Any:
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        try:
            return float(text) if "." in text else int(text)
        except ValueError:
            return text
    return value


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\uff08", "(").replace("\uff09", ")").replace("\uff0f", "/")
    return re.sub(r"[\s\r\n\t:_：,，;；]+", "", text)


def _header_matches_alias(header: str, canonical: str) -> bool:
    header_key = _normalise_header(header)
    if not header_key:
        return False
    for alias in HEADER_ALIASES.get(canonical, (canonical,)):
        alias_key = _normalise_header(alias)
        if not alias_key:
            continue
        if header_key == alias_key:
            return True
        # Intake workbooks often add qualifiers such as "厂房/谱表",
        # "后补报告号" or "方钢截面(mm)". Treat those as the same field,
        # but do not allow a very short alias like "类型" to match unrelated
        # columns accidentally.
        if len(alias_key) >= 3 and alias_key in header_key:
            return True
        if len(header_key) >= 3 and header_key in alias_key:
            return True
    return False


def _alias_lookup(row: dict[str, Any], canonical: str) -> Any:
    normalised = {_normalise_header(key): value for key, value in row.items()}
    for alias in HEADER_ALIASES.get(canonical, (canonical,)):
        key = _normalise_header(alias)
        if key in normalised and normalised[key] not in (None, ""):
            return normalised[key]
    for header, value in row.items():
        if value not in (None, "") and _header_matches_alias(str(header), canonical):
            return value
    return None


def read_key_value_intake(path: Path | str) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.lower() == ".json":
        return json.loads(path.read_text(encoding="utf-8-sig"))
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return {row[0].strip(): _coerce(row[1]) for row in csv.reader(handle) if len(row) >= 2 and row[0].strip()}
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read intake Excel files") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    payload: dict[str, Any] = {}
    for key, value in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if key is None:
            continue
        payload[str(key).strip()] = _coerce(value)
    return payload


def _parse_meter(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    match = re.search(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", text)
    return float(match.group(0)) if match else None


def _parse_meter_candidates(value: Any) -> list[float]:
    if value is None:
        return []
    if isinstance(value, (int, float)):
        return [float(value)]
    values: list[float] = []
    for match in re.finditer(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", str(value)):
        try:
            values.append(float(match.group(0)))
        except ValueError:
            continue
    return values


def _default_project_code(path: Path) -> str:
    match = re.match(r"\s*(\d{4})", path.stem)
    if match:
        return match.group(1)
    first_token = re.split(r"[\s_\-]+", path.stem.strip(), maxsplit=1)[0]
    return first_token or path.stem


def _fallback_report_number(support_type: Any, serial: Any, row_number: int) -> str:
    left = str(support_type or "JOB").strip() or "JOB"
    right = str(serial or row_number).strip() or str(row_number)
    return f"{left}-{right}"


def _provisional_intake_id(source_file: Path, sheet_name: str, serial: Any, row_number: int) -> str:
    source_token = re.sub(r"\W+", "_", source_file.stem, flags=re.UNICODE).strip("_") or "intake"
    sheet_token = re.sub(r"\W+", "_", sheet_name, flags=re.UNICODE).strip("_") or "sheet"
    serial_token = re.sub(r"\W+", "_", str(serial or row_number), flags=re.UNICODE).strip("_") or str(row_number)
    return f"{source_token}_{sheet_token}_row_{serial_token}"


def _normalise_square_section(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip().upper()
    match = re.search(r"(\d+(?:\.\d+)?)\s*[-X×脳*]\s*(\d+(?:\.\d+)?)\s*[-X×脳*]\s*(\d+(?:\.\d+)?)", text)
    if not match:
        return text
    first, second, third = (float(item) for item in match.groups())
    if abs(first - second) > 1e-9:
        return text.replace("X", "-").replace("×", "-").replace("脳", "-").replace("*", "-")
    return f"{first:g}-{second:g}-{third:g}"


def _normalise_square_section(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    match = re.search(r"(\d+(?:\.\d+)?)\s*[-*xX×＊]\s*(\d+(?:\.\d+)?)\s*[-*xX×＊]\s*(\d+(?:\.\d+)?)", text)
    if not match:
        return text
    first, second, third = (float(item) for item in match.groups())
    if abs(first - second) > 1e-9:
        return re.sub(r"[-*xX×＊]+", "-", text)
    return f"{first:g}-{second:g}-{third:g}"


def _extract_allowed_square_sections_from_text(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    found: list[str] = []
    for match in re.finditer(
        r"(\d+(?:\.\d+)?)\s*[-*xX×＊]\s*(\d+(?:\.\d+)?)\s*[-*xX×＊]\s*(\d+(?:\.\d+)?)",
        str(value),
    ):
        first, second, third = (float(item) for item in match.groups())
        if abs(first - second) > 1e-9:
            continue
        normalized = f"{first:g}-{second:g}-{third:g}"
        if normalized not in found:
            found.append(normalized)
    return found


def _find_allowed_square_sections_from_sheet_items(
    sheet_items: list[tuple[str, list[tuple[Any, ...] | list[Any]]]],
) -> dict[str, Any]:
    best: dict[str, Any] = {
        "allowed_square_section_ids": [],
        "allowed_square_section_source_ref": None,
        "allowed_square_section_status": "not_found",
    }
    best_score = -1
    for sheet_name, sheet_values in sheet_items:
        sheet_score = 3 if any(token in str(sheet_name) for token in ("计算说明", "说明")) else 0
        for row_index, row in enumerate(sheet_values, start=1):
            for col_index, value in enumerate(row, start=1):
                sections = _extract_allowed_square_sections_from_text(value)
                if not sections:
                    continue
                text = str(value)
                score = sheet_score + sum(
                    2 for token in ("计算说明", "说明", "方钢", "方刚", "抗震", "最小方钢", "满足") if token in text
                )
                if score > best_score or (score == best_score and len(sections) > len(best["allowed_square_section_ids"])):
                    best_score = score
                    best = {
                        "allowed_square_section_ids": sections,
                        "allowed_square_section_source_ref": f"{sheet_name}!R{row_index}C{col_index}",
                        "allowed_square_section_status": "provided_by_intake_calculation_notes",
                    }
    return best


def _looks_like_support_type(value: Any) -> bool:
    text = str(value or "").strip().upper()
    return bool(re.search(r"(?<![A-Z0-9])S\d+[A-Z]?(?![A-Z0-9])", text)) or "\u94a2\u5e73\u53f0" in text


def _looks_like_data_row(row: dict[str, Any]) -> bool:
    support_type = _alias_lookup(row, "support_type")
    if not _looks_like_support_type(support_type):
        return False
    score = 0
    for canonical, weight in (
        ("description", 2),
        ("building", 2),
        ("elevation", 2),
        ("support_spacing_m", 1),
        ("support_height_m", 1),
        ("report_number", 1),
        ("support_id", 1),
    ):
        if _alias_lookup(row, canonical) not in (None, ""):
            score += weight
    return score >= 4


def determine_analysis_method(row: dict[str, Any], payload: dict[str, Any] | None = None) -> str:
    explicit = _alias_lookup(row, "analysis_method")
    if explicit:
        text = str(explicit).strip().lower()
        if "static" in text or "\u9759\u529b" in text:
            return "static"
        if "spectrum" in text or "\u53cd\u5e94\u8c31" in text:
            return "response_spectrum"

    combined = " ".join(str(value) for value in [*row.values(), *(payload or {}).values()] if value is not None)
    return "static" if STATIC_METHOD_KEYWORD in combined else "response_spectrum"


def _normalise_table_row(row: dict[str, Any], *, source_file: Path, sheet_name: str, row_number: int) -> dict[str, Any]:
    serial = _alias_lookup(row, "serial")
    if serial is None:
        serial = row_number
    if not _looks_like_data_row(row):
        raise ValueError("tabular intake row does not look like a calculation task")
    support_type = _alias_lookup(row, "support_type") or "S2"
    building = _alias_lookup(row, "building") or "UNSPECIFIED"
    building = str(building).strip()
    area = str(_alias_lookup(row, "area") or building).strip()
    elevation_raw = _alias_lookup(row, "elevation")
    elevation_candidates = _parse_meter_candidates(elevation_raw)
    elevation = elevation_candidates[0] if elevation_candidates else None
    spacing = _parse_meter(_alias_lookup(row, "support_spacing_m"))
    height = _parse_meter(_alias_lookup(row, "support_height_m"))
    report_number = _alias_lookup(row, "report_number")
    provisional_id = _provisional_intake_id(source_file, sheet_name, serial, row_number)
    description = _alias_lookup(row, "description") or "Created from uploaded intake workbook"
    square_section_spec = _normalise_square_section(_alias_lookup(row, "square_section_spec"))
    cantilever_evaluation_mode = _alias_lookup(row, "cantilever_evaluation_mode")
    payload = {
        "project_code": _alias_lookup(row, "project_code") or _default_project_code(source_file),
        "building": str(building),
        "area": str(area),
        "elevation": elevation if elevation is not None else 0.0,
        "elevation_raw": elevation_raw,
        "elevation_candidates": elevation_candidates,
        "damping_ratio": _alias_lookup(row, "damping_ratio") or 0.1,
        "material": _alias_lookup(row, "material") or "Q355",
        "support_type": str(support_type),
        "support_spacing_m": spacing,
        "support_height_m": height,
        "description": description,
        "report_number": report_number,
        "calculation_batch": report_number,
        "intake_order_id": report_number or provisional_id,
        "provisional_intake_id": provisional_id,
        "intake_identity_status": "formal_report_number_provided" if report_number else "provisional_before_report_number",
        "intake_serial": serial,
        "support_id": _alias_lookup(row, "support_id"),
        "intake_row_number": row_number,
        "intake_sheet": sheet_name,
        "raw_intake_row": row,
    }
    if square_section_spec:
        payload["square_section_spec"] = str(square_section_spec)
        payload["square_section_source"] = "intake_column_I_embedded_plate_header_is_square_section"
    else:
        payload["square_section_source"] = "pending_manual_selection_after_calculation"
    if cantilever_evaluation_mode:
        payload["cantilever_evaluation_mode"] = str(cantilever_evaluation_mode)
    payload["analysis_method"] = determine_analysis_method(row, payload)
    coerced = {key: _coerce(value) for key, value in payload.items() if value is not None}
    text_keys = {
        "project_code",
        "building",
        "area",
        "support_type",
        "description",
        "report_number",
        "calculation_batch",
        "intake_order_id",
        "provisional_intake_id",
        "intake_identity_status",
        "support_id",
        "intake_sheet",
        "square_section_spec",
        "square_section_source",
        "cantilever_evaluation_mode",
        "analysis_method",
    }
    for key in text_keys:
        if key in coerced and coerced[key] is not None:
            coerced[key] = str(coerced[key])
    return coerced


def _looks_like_header(headers: list[str]) -> bool:
    matched_fields = {
        canonical
        for canonical in HEADER_ALIASES
        if any(_header_matches_alias(header, canonical) for header in headers)
    }
    score = 0
    for canonical, weight in (
        ("support_type", 2),
        ("description", 2),
        ("building", 2),
        ("elevation", 2),
        ("report_number", 1),
        ("serial", 1),
        ("square_section_spec", 1),
        ("support_spacing_m", 1),
        ("support_height_m", 1),
    ):
        if canonical in matched_fields:
            score += weight
    return score >= 5 and bool({"support_type", "description", "building", "elevation"} & matched_fields)


def _strip_xml_tags(value: str) -> str:
    return html.unescape(re.sub(r"<[^>]+>", "", value))


def _read_xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        text = archive.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
    except KeyError:
        return []
    strings: list[str] = []
    for item in re.findall(r"<(?:\w+:)?si\b[^>]*>(.*?)</(?:\w+:)?si>", text, flags=re.S):
        parts = re.findall(r"<(?:\w+:)?t\b[^>]*>(.*?)</(?:\w+:)?t>", item, flags=re.S)
        strings.append("".join(_strip_xml_tags(part) for part in parts))
    return strings


def _xlsx_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    try:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8", errors="replace")
    except KeyError:
        return []
    relationships: dict[str, str] = {}
    try:
        rels_xml = archive.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
        for attrs in re.findall(r"<(?:\w+:)?Relationship\b([^>]*)/?>", rels_xml, flags=re.S):
            rid_match = re.search(r'\bId="([^"]+)"', attrs)
            target_match = re.search(r'\bTarget="([^"]+)"', attrs)
            if rid_match and target_match:
                target = target_match.group(1).replace("\\", "/")
                if target.startswith("/"):
                    target = target.lstrip("/")
                elif not target.startswith("xl/"):
                    target = "xl/" + target.lstrip("/")
                relationships[rid_match.group(1)] = target
    except KeyError:
        pass
    sheets: list[tuple[str, str]] = []
    for attrs in re.findall(r"<(?:\w+:)?sheet\b([^>]*)/?>", workbook_xml, flags=re.S):
        name_match = re.search(r'\bname="([^"]*)"', attrs)
        rid_match = re.search(r'\b(?:\w+:)?id="([^"]+)"', attrs)
        if not name_match:
            continue
        sheet_name = html.unescape(name_match.group(1))
        target = relationships.get(rid_match.group(1) if rid_match else "")
        if target:
            sheets.append((sheet_name, target))
    if sheets:
        return sheets
    return [
        (f"sheet{index}", name)
        for index, name in enumerate(sorted(item for item in archive.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", item)), start=1)
    ]


def _column_index(cell_ref: str) -> int:
    letters = re.match(r"([A-Z]+)", cell_ref.upper())
    if not letters:
        return 0
    value = 0
    for char in letters.group(1):
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value - 1


def _xlsx_cell_value(cell_xml: str, cell_type: str, shared_strings: list[str]) -> Any:
    inline = re.search(r"<(?:\w+:)?is\b[^>]*>(.*?)</(?:\w+:)?is>", cell_xml, flags=re.S)
    if inline:
        parts = re.findall(r"<(?:\w+:)?t\b[^>]*>(.*?)</(?:\w+:)?t>", inline.group(1), flags=re.S)
        return _coerce("".join(_strip_xml_tags(part) for part in parts))
    value_match = re.search(r"<(?:\w+:)?v\b[^>]*>(.*?)</(?:\w+:)?v>", cell_xml, flags=re.S)
    raw = _strip_xml_tags(value_match.group(1)) if value_match else ""
    if cell_type == "s":
        try:
            return shared_strings[int(float(raw))]
        except (ValueError, IndexError):
            return raw
    if cell_type == "b":
        return raw in {"1", "true", "TRUE"}
    return _coerce(raw)


def _iter_xlsx_cells(row_xml: str):
    cell_pattern = re.compile(
        r"<(?:\w+:)?c\b([^>]*)/>|<(?:\w+:)?c\b([^>]*)>(.*?)</(?:\w+:)?c>",
        flags=re.S,
    )
    for match in cell_pattern.finditer(row_xml):
        attrs = match.group(1) if match.group(1) is not None else match.group(2)
        body = match.group(3) or ""
        yield attrs, body


def _read_xlsx_rows_without_expat(path: Path) -> list[tuple[str, list[list[Any]]]]:
    """Read enough XLSX structure for intake sheets when packaged expat is unavailable."""
    sheets: list[tuple[str, list[list[Any]]]] = []
    with zipfile.ZipFile(path, "r") as archive:
        shared_strings = _read_xlsx_shared_strings(archive)
        for sheet_name, target in _xlsx_sheet_targets(archive):
            try:
                sheet_xml = archive.read(target).decode("utf-8", errors="replace")
            except KeyError:
                continue
            rows: list[list[Any]] = []
            row_pattern = re.compile(
                r"<(?:\w+:)?row\b([^>]*)/>|<(?:\w+:)?row\b([^>]*)>(.*?)</(?:\w+:)?row>",
                flags=re.S,
            )
            for row_match in row_pattern.finditer(sheet_xml):
                row_attrs = row_match.group(1) if row_match.group(1) is not None else row_match.group(2) or ""
                row_xml = row_match.group(3) or ""
                row_ref = re.search(r'\br="(\d+)"', row_attrs)
                target_index = int(row_ref.group(1)) - 1 if row_ref else len(rows)
                while len(rows) < target_index:
                    rows.append([])
                cells: dict[int, Any] = {}
                for attrs, body in _iter_xlsx_cells(row_xml):
                    ref_match = re.search(r'\br="([^"]+)"', attrs)
                    type_match = re.search(r'\bt="([^"]+)"', attrs)
                    index = _column_index(ref_match.group(1)) if ref_match else len(cells)
                    cells[index] = _xlsx_cell_value(body, type_match.group(1) if type_match else "", shared_strings)
                if cells:
                    max_index = max(cells)
                    rows.append([cells.get(index) for index in range(max_index + 1)])
                elif len(rows) == target_index:
                    rows.append([])
            sheets.append((sheet_name, rows))
    return sheets


def _openpyxl_unavailable_due_to_expat(exc: Exception) -> bool:
    message = str(exc)
    return "No module named expat" in message or "SimpleXMLTreeBuilder" in message


def read_tabular_intake_rows(path: Path | str) -> list[dict[str, Any]]:
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read intake Excel files") from exc
    rows: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet_items = [(sheet.title, list(sheet.iter_rows(values_only=True))) for sheet in workbook.worksheets]
    except Exception as exc:
        if not _openpyxl_unavailable_due_to_expat(exc):
            raise
        sheet_items = _read_xlsx_rows_without_expat(path)
    allowed_square_sections = _find_allowed_square_sections_from_sheet_items(sheet_items)
    for sheet_name, sheet_values in sheet_items:
        for header_index, header_values in enumerate(sheet_values, start=1):
            headers = [str(value).strip() if value is not None else "" for value in header_values]
            if not _looks_like_header(headers):
                continue
            for row_number, values in enumerate(sheet_values[header_index:], start=header_index + 1):
                row = {
                    header: _coerce(value)
                    for header, value in zip(headers, values)
                    if header and value not in (None, "")
                }
                if not row:
                    continue
                try:
                    normalised = _normalise_table_row(
                        row,
                        source_file=path,
                        sheet_name=sheet_name,
                        row_number=row_number,
                    )
                    if allowed_square_sections["allowed_square_section_ids"]:
                        normalised.update(allowed_square_sections)
                    rows.append(normalised)
                except ValueError:
                    continue
            if rows:
                return rows
    return rows


def select_tabular_intake_row(
    path: Path | str,
    *,
    intake_order_id: str | None = None,
    report_number: str | None = None,
    row_number: int | None = None,
) -> dict[str, Any]:
    rows = read_tabular_intake_rows(path)
    if not rows:
        raise ValueError("No tabular intake rows were found.")
    wanted_id = report_number or intake_order_id
    if wanted_id:
        for row in rows:
            keys = (row.get("report_number"), row.get("calculation_batch"), row.get("intake_order_id"))
            if any(str(value) == str(wanted_id) for value in keys if value is not None):
                return row
        raise ValueError(f"Intake report/calculation batch was not found: {wanted_id}")
    if row_number:
        for row in rows:
            if int(row.get("intake_row_number", -1)) == int(row_number):
                return row
            try:
                serial = int(float(str(row.get("intake_serial"))))
            except (TypeError, ValueError):
                serial = None
            if serial == int(row_number):
                return row
        raise ValueError(f"Intake row was not found: {row_number}")
    return rows[0]


def validate_intake_payload(payload: dict[str, Any]) -> dict[str, Any]:
    missing = [field for field in REQUIRED_FIELDS if payload.get(field) in (None, "")]
    return {
        "status": "fail" if missing else "pass",
        "missing_fields": missing,
        "field_count": len(payload),
    }


def read_and_validate_intake(
    path: Path | str,
    *,
    intake_order_id: str | None = None,
    report_number: str | None = None,
    row_number: int | None = None,
) -> dict[str, Any]:
    try:
        payload = read_key_value_intake(path)
        validation = validate_intake_payload(payload)
        intake_format = "key_value"
    except Exception as exc:
        if Path(path).suffix.lower() not in {".xlsx", ".xlsm"}:
            raise
        payload = {"key_value_read_error": str(exc)}
        validation = {"status": "fail", "missing_fields": REQUIRED_FIELDS, "field_count": 0}
        intake_format = "key_value_failed"
    if validation["status"] != "pass":
        try:
            payload = select_tabular_intake_row(
                path,
                intake_order_id=intake_order_id,
                report_number=report_number,
                row_number=row_number,
            )
            validation = validate_intake_payload(payload)
            intake_format = "tabular"
        except Exception:
            pass
    return {
        "payload": payload,
        "validation": validation,
        "source_file": str(path),
        "intake_format": intake_format,
    }
