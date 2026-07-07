from __future__ import annotations

import argparse
import json
import sys
import tempfile
from pathlib import Path
from typing import Any


def _add_segmented_section(sheet: Any, label: str, scale: float) -> None:
    sheet.append([label])
    for frequency, base in ((0.1, 0.015), (3.927, 0.733), (4.094, 0.658), (5.482, 1.514), (100.0, 0.477)):
        sheet.append([None, frequency, base * scale, base * scale, base * scale, base * scale, base * scale, base * scale])


def _add_precision_control_sheet(workbook: Any) -> None:
    sheet = workbook.create_sheet("Precision")
    values = [50, 0, 33, 0, 15, 0, 3, 0, 0]
    for group_index in range(1, 9):
        row0 = 5 + ((group_index - 1) % 4) * 19
        col0 = 1 + ((group_index - 1) // 4) * 11
        for offset, value in enumerate(values):
            sheet.cell(row0 + offset, col0, value)


def _write_smoke_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    review = workbook.active
    review.title = "Review"
    ansys_lines = [
        "!SL-1(XY) 7%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,100.000",
        "SV, 0.07,  0.015,  0.733,  1.514,  0.182",
        "!SL-1(Z) 7%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,100.000",
        "SV, 0.07,  0.009,  0.502,  1.231,  0.117",
        "!SL-2(XY) 10%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000",
        "SV, 0.10,  0.043,  0.733,  1.514,  1.737,  0.477",
        "!SL-2(Z) 10%  Envelop:(NR_1818,8.5)",
        "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000",
        "SV, 0.10,  0.024,  0.451,  1.231,  1.360,  0.337",
    ]
    for row_number, line in enumerate(ansys_lines, start=1):
        review.cell(row_number, 13, line)

    segmented = workbook.create_sheet("NR_1818")
    for label, scale in (
        ("8.45x", 1.0),
        ("8.45y", 1.1),
        ("8.45z", 0.8),
        ("8.45X", 1.2),
        ("8.45Y", 1.3),
        ("8.45Z", 0.9),
    ):
        _add_segmented_section(segmented, label, scale)
    _add_precision_control_sheet(workbook)
    workbook.save(path)
    workbook.close()


def _assert_absent(package_dir: Path, relative_paths: list[str]) -> list[str]:
    failures: list[str] = []
    for relative in relative_paths:
        candidate = package_dir / relative
        if candidate.exists():
            failures.append(relative)
    return failures


def _runtime_xml_files(package_dir: Path) -> dict[str, Any]:
    internal = package_dir / "runtime" / "CableTrayAI_Server" / "_internal"
    required = ["pyexpat.pyd", "_elementtree.pyd", "libexpat.dll"]
    return {
        "internal": str(internal),
        "required": required,
        "missing": [name for name in required if not (internal / name).exists()],
    }


def _run_no_expat_spectrum_smoke(package_dir: Path) -> dict[str, Any]:
    sys.path.insert(0, str(package_dir))

    import openpyxl

    with tempfile.TemporaryDirectory(prefix="cabletrayai_package_gate_") as temp_name:
        temp = Path(temp_name)
        workbook = temp / "spectrum.xlsx"
        _write_smoke_workbook(workbook)

        def fail_load_workbook(*_args: Any, **_kwargs: Any) -> None:
            raise ImportError("No module named expat; use SimpleXMLTreeBuilder instead")

        openpyxl.load_workbook = fail_load_workbook

        from core.spectra.response_spectrum_writer import write_segmented_response_spectrum_mac

        audit = write_segmented_response_spectrum_mac(
            workbook,
            temp / "out",
            project_code="1818",
            building="NR",
            elevation=8.5,
        )
        sl2 = (temp / "out" / "ansys_spectrum_sl2.mac").read_text(encoding="utf-8")
        expected = "FREQ    ,  0.100,  3.927,  5.482,  5.715,100.000"
        if audit.get("active_workbook_column_m_comparison", {}).get("status") != "pass":
            raise AssertionError("No-expat spectrum smoke did not match active column M ANSYS Format")
        if expected not in sl2:
            raise AssertionError("No-expat spectrum smoke did not write the calibrated SL-2 spectrum block")
        return {
            "status": "pass",
            "formal_spectrum_source_mode": audit.get("formal_spectrum_source_mode"),
            "comparison_status": audit.get("active_workbook_column_m_comparison", {}).get("status"),
        }


def main() -> int:
    parser = argparse.ArgumentParser(description="CableTrayAI deployment package gate")
    parser.add_argument("--package-dir", required=True)
    args = parser.parse_args()

    package_dir = Path(args.package_dir).resolve()
    result: dict[str, Any] = {
        "status": "pass",
        "package_dir": str(package_dir),
        "checks": {},
    }
    forbidden = _assert_absent(
        package_dir,
        [
            "jobs",
            "uploads",
            "outputs",
            "logs",
            "docs/paper",
            "runtime/auth_sessions.json",
            "config/ansys.local.toml",
            "config/access_control.local.json",
            "core/spectra/interpolation.py",
        ],
    )
    result["checks"]["forbidden_paths_absent"] = {"status": "pass" if not forbidden else "fail", "found": forbidden}
    runtime_xml = _runtime_xml_files(package_dir)
    result["checks"]["runtime_xml_support_files"] = {
        "status": "pass" if not runtime_xml["missing"] else "fail",
        **runtime_xml,
    }
    if forbidden or runtime_xml["missing"]:
        result["status"] = "fail"
    else:
        try:
            result["checks"]["no_expat_spectrum_smoke"] = _run_no_expat_spectrum_smoke(package_dir)
        except Exception as exc:
            result["status"] = "fail"
            result["checks"]["no_expat_spectrum_smoke"] = {"status": "fail", "error": str(exc)}

    docs_dir = package_dir / "docs"
    docs_dir.mkdir(parents=True, exist_ok=True)
    (docs_dir / "deployment_package_gate.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
